#Name : William Karanja
# Date : 23/02/2026
# Program to show use of objects and classes


import tkinter as tk
from tkinter import ttk, messagebox

class SchoolManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("School Management System")
        self.root.geometry("1000x600")
        self.root.config(bg="#f0f0f0")

        self.students = []

        # ===== TITLE =====
        title = tk.Label(root, text="SCHOOL MANAGEMENT SYSTEM",
                         font=("Arial", 20, "bold"),
                         bg="#1e3d59", fg="white")
        title.pack(fill="x")

        # ===== INPUT FRAME =====
        frame = tk.Frame(root, bg="#f0f0f0")
        frame.pack(pady=10)

        tk.Label(frame, text="Student ID").grid(row=0, column=0, padx=5, pady=5)
        self.id_entry = tk.Entry(frame)
        self.id_entry.grid(row=0, column=1)

        tk.Label(frame, text="Student Name").grid(row=1, column=0, padx=5, pady=5)
        self.name_entry = tk.Entry(frame)
        self.name_entry.grid(row=1, column=1)

        tk.Label(frame, text="Course").grid(row=2, column=0, padx=5, pady=5)
        self.course_entry = tk.Entry(frame)
        self.course_entry.grid(row=2, column=1)

        tk.Label(frame, text="Contact").grid(row=3, column=0, padx=5, pady=5)
        self.contact_entry = tk.Entry(frame)
        self.contact_entry.grid(row=3, column=1)

        tk.Button(frame, text="Register Student",
                  command=self.register_student,
                  bg="green", fg="white").grid(row=4, columnspan=2, pady=10)

        # ===== TABLE =====
        self.tree = ttk.Treeview(root,
                                 columns=("ID", "Name", "Course", "Contact", "Grade"),
                                 show="headings")

        for col in ("ID", "Name", "Course", "Contact", "Grade"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)

        self.tree.pack(fill="both", expand=True, padx=20, pady=10)

        # ===== UPDATE FRAME =====
        update_frame = tk.Frame(root, bg="#f0f0f0")
        update_frame.pack(pady=10)

        tk.Label(update_frame, text="New Course").grid(row=0, column=0, padx=5)
        self.new_course_entry = tk.Entry(update_frame)
        self.new_course_entry.grid(row=0, column=1)

        tk.Label(update_frame, text="Grade").grid(row=0, column=2, padx=5)
        self.grade_entry = tk.Entry(update_frame)
        self.grade_entry.grid(row=0, column=3)

        tk.Button(update_frame, text="Assign Course & Grade",
                  command=self.assign_course_grade,
                  bg="blue", fg="white").grid(row=0, column=4, padx=10)

    # ===== FUNCTIONS =====

    def register_student(self):
        student_id = self.id_entry.get()
        name = self.name_entry.get()
        course = self.course_entry.get()
        contact = self.contact_entry.get()

        if not student_id or not name or not course or not contact:
            messagebox.showerror("Error", "All fields are required!")
            return

        self.tree.insert("", "end",
                         values=(student_id, name, course, contact, ""))

        self.students.append({
            "id": student_id,
            "name": name,
            "course": course,
            "contact": contact,
            "grade": ""
        })

        self.clear_entries()
        messagebox.showinfo("Success", "Student Registered Successfully!")

    def assign_course_grade(self):
        selected = self.tree.selection()

        if not selected:
            messagebox.showerror("Error", "Select a student first!")
            return

        new_course = self.new_course_entry.get()
        grade = self.grade_entry.get()

        if not new_course or not grade:
            messagebox.showerror("Error", "Enter new course and grade!")
            return

        for item in selected:
            values = self.tree.item(item, "values")
            updated_values = (values[0], values[1],
                              new_course, values[3], grade)
            self.tree.item(item, values=updated_values)

        self.new_course_entry.delete(0, tk.END)
        self.grade_entry.delete(0, tk.END)

        messagebox.showinfo("Updated", "Course and Grade Assigned!")

    def clear_entries(self):
        self.id_entry.delete(0, tk.END)
        self.name_entry.delete(0, tk.END)
        self.course_entry.delete(0, tk.END)
        self.contact_entry.delete(0, tk.END)


# ===== RUN PROGRAM =====
if __name__ == "__main__":
    root = tk.Tk()
    app = SchoolManagementSystem(root)
    root.mainloop()




import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "students.xlsx"

class SchoolManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("School Management System")
        self.root.geometry("1000x600")

        # Create Excel file if not exists
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Name", "Course", "Contact", "Grade"])
            wb.save(FILE_NAME)

        # ===== INPUT FRAME =====
        frame = tk.Frame(root)
        frame.pack(pady=10)

        tk.Label(frame, text="Student ID").grid(row=0, column=0)
        self.id_entry = tk.Entry(frame)
        self.id_entry.grid(row=0, column=1)

        tk.Label(frame, text="Student Name").grid(row=1, column=0)
        self.name_entry = tk.Entry(frame)
        self.name_entry.grid(row=1, column=1)

        tk.Label(frame, text="Course").grid(row=2, column=0)
        self.course_entry = tk.Entry(frame)
        self.course_entry.grid(row=2, column=1)

        tk.Label(frame, text="Contact").grid(row=3, column=0)
        self.contact_entry = tk.Entry(frame)
        self.contact_entry.grid(row=3, column=1)

        tk.Button(frame, text="Register Student",
                  command=self.register_student,
                  bg="green", fg="white").grid(row=4, columnspan=2, pady=10)

        # ===== TABLE =====
        self.tree = ttk.Treeview(root,
                                 columns=("ID", "Name", "Course", "Contact", "Grade"),
                                 show="headings")

        for col in ("ID", "Name", "Course", "Contact", "Grade"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)

        self.tree.pack(fill="both", expand=True, padx=20, pady=10)

        # ===== UPDATE FRAME =====
        update_frame = tk.Frame(root)
        update_frame.pack(pady=10)

        tk.Label(update_frame, text="New Course").grid(row=0, column=0)
        self.new_course_entry = tk.Entry(update_frame)
        self.new_course_entry.grid(row=0, column=1)

        tk.Label(update_frame, text="Grade").grid(row=0, column=2)
        self.grade_entry = tk.Entry(update_frame)
        self.grade_entry.grid(row=0, column=3)

        tk.Button(update_frame, text="Assign Course & Grade",
                  command=self.assign_course_grade,
                  bg="blue", fg="white").grid(row=0, column=4)

        self.load_data()

    # ===== REGISTER STUDENT =====
    def register_student(self):
        student_id = self.id_entry.get()
        name = self.name_entry.get()
        course = self.course_entry.get()
        contact = self.contact_entry.get()

        if not student_id or not name or not course or not contact:
            messagebox.showerror("Error", "All fields required!")
            return

        # Save to Excel
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append([student_id, name, course, contact, ""])
        wb.save(FILE_NAME)

        self.tree.insert("", "end",
                         values=(student_id, name, course, contact, ""))

        self.clear_entries()
        messagebox.showinfo("Success", "Student Saved to Excel!")

    # ===== ASSIGN COURSE & GRADE =====
    def assign_course_grade(self):
        selected = self.tree.selection()

        if not selected:
            messagebox.showerror("Error", "Select student first!")
            return

        new_course = self.new_course_entry.get()
        grade = self.grade_entry.get()

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        for item in selected:
            values = self.tree.item(item, "values")
            student_id = values[0]

            for row in ws.iter_rows(min_row=2):
                if row[0].value == student_id:
                    row[2].value = new_course
                    row[4].value = grade

            self.tree.item(item,
                           values=(values[0], values[1],
                                   new_course, values[3], grade))

        wb.save(FILE_NAME)

        self.new_course_entry.delete(0, tk.END)
        self.grade_entry.delete(0, tk.END)

        messagebox.showinfo("Updated", "Excel Updated Successfully!")

    # ===== LOAD DATA FROM EXCEL =====
    def load_data(self):
        wb = load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)

    def clear_entries(self):
        self.id_entry.delete(0, tk.END)
        self.name_entry.delete(0, tk.END)
        self.course_entry.delete(0, tk.END)
        self.contact_entry.delete(0, tk.END)


# ===== RUN =====
if __name__ == "__main__":
    root = tk.Tk()
    app = SchoolManagementSystem(root)
    root.mainloop() 

  # ===== DELETE STUDENT =====
def delete_student(self):
    selected = self.tree.selection()

    if not selected:
        messagebox.showerror("Error", "Select a student to delete!")
        return

    confirm = messagebox.askyesno("Confirm Delete",
                                   "Are you sure you want to delete this student?")

    if not confirm:
        return

    for item in selected:
        values = self.tree.item(item, "values")
        student_id = values[0]

        # Remove from internal list
        self.students = [student for student in self.students
                         if student["id"] != student_id]

        # Remove from table
        self.tree.delete(item)

    messagebox.showinfo("Deleted", "Student deleted successfully!")  